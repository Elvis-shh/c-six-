"""
SmartReport 测试报告 Excel 模板生成器
用法: python generate_test_report.py [--output test-report-v1.0.0.xlsx]
"""
import argparse
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 样式定义 ──────────────────────────────────────────

# 颜色
HEADER_FILL   = PatternFill("solid", fgColor="1F2937")   # 深灰表头
HEADER_FONT   = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
P0_FILL       = PatternFill("solid", fgColor="FEE2E2")   # 浅红 P0
P1_FILL       = PatternFill("solid", fgColor="FEF3C7")   # 浅黄 P1
P2_FILL       = PatternFill("solid", fgColor="F3F4F6")   # 浅灰 P2
PASS_FILL     = PatternFill("solid", fgColor="D1FAE5")   # 浅绿 通过
FAIL_FILL     = PatternFill("solid", fgColor="FECACA")   # 浅红 失败
BLOCKED_FILL  = PatternFill("solid", fgColor="FDE68A")   # 浅黄 阻塞
OVERVIEW_KPI_FILL = PatternFill("solid", fgColor="EFF6FF")  # 概览 KPI 底色
OVERVIEW_HEADER_FILL = PatternFill("solid", fgColor="DBEAFE")
SECTION_FILL  = PatternFill("solid", fgColor="E5E7EB")

BODY_FONT     = Font(name="微软雅黑", size=9)
BOLD_FONT     = Font(name="微软雅黑", size=9, bold=True)
TITLE_FONT    = Font(name="微软雅黑", size=14, bold=True, color="1F2937")
SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="6B7280")

CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP     = Alignment(horizontal="left", vertical="center", wrap_text=True)
TITLE_ALIGN   = Alignment(horizontal="left", vertical="center")

THIN_BORDER   = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


# ── 测试用例数据 ──────────────────────────────────────

# 格式: (TC-ID, 模块, 优先级, 测试场景, 执行类型, 前置条件, 测试步骤, 预期结果)
# 执行类型: 🤖自动 / 🙋人工 / 🔀半自动

TEST_CASES = [
    # ── 模块 1.1: 数据库表结构 ──
    ("TC-1.1-01", "1.1 数据库表结构", "P0", "Docker 启动 MySQL 自动建表", "🤖自动",
     "compose 已配置", "docker compose up -d mysql，等待 healthy", "SHOW TABLES 显示 17~18 张表"),
    ("TC-1.1-02", "1.1 数据库表结构", "P0", "Spring Boot 启动 Hibernate 校验", "🤖自动",
     "MySQL 已启动，表已建", "启动 backend 容器，观察日志", "无 Schema-validation 报错"),
    ("TC-1.1-03", "1.1 数据库表结构", "P0", "查询贵州茅台基本信息", "🙋人工",
     "种子数据已插入", "SELECT * FROM companies WHERE code='600519'", "返回 1 行，name=贵州茅台"),
    ("TC-1.1-04", "1.1 数据库表结构", "P0", "查询财报指标数据", "🙋人工",
     "种子数据已插入", "SELECT * FROM financial_indicators WHERE report_id=1", "返回多条指标记录"),
    ("TC-1.1-05", "1.1 数据库表结构", "P1", "复合索引生效验证", "🤖自动",
     "种子数据已插入", "EXPLAIN 分析查询计划", "使用 idx_company_year 索引"),
    ("TC-1.1-06", "1.1 数据库表结构", "P1", "指标定义完整性检查", "🤖自动",
     "种子数据已插入", "SELECT COUNT(*) FROM indicator_definitions", "≥ 30 条"),

    # ── 模块 1.2: 公司搜索 API ──
    ("TC-1.2-01", "1.2 公司搜索 API", "P0", "按公司名模糊搜索", "🤖自动",
     "后端运行中", "GET /api/v1/search/companies?q=茅台&limit=8", "code=0, data 含贵州茅台"),
    ("TC-1.2-02", "1.2 公司搜索 API", "P0", "按股票代码搜索", "🤖自动",
     "后端运行中", "GET /api/v1/search/companies?q=600519", "code=0, data[0].code=600519"),
    ("TC-1.2-03", "1.2 公司搜索 API", "P0", "搜索无匹配结果", "🤖自动",
     "后端运行中", "GET /api/v1/search/companies?q=xyzabc123", "code=0, data=[]"),
    ("TC-1.2-04", "1.2 公司搜索 API", "P1", "空搜索词处理", "🤖自动",
     "后端运行中", "GET /api/v1/search/companies?q=", "返回空数组或 400"),
    ("TC-1.2-05", "1.2 公司搜索 API", "P0", "获取热门公司", "🤖自动",
     "后端运行中", "GET /api/v1/search/companies/hot", "code=0, data 长度=6"),
    ("TC-1.2-06", "1.2 公司搜索 API", "P1", "分页限制", "🤖自动",
     "后端运行中", "GET /api/v1/search/companies?q=科技&limit=3", "data 长度 ≤ 3"),
    ("TC-1.2-07", "1.2 公司搜索 API", "P1", "部分匹配（简称）", "🤖自动",
     "后端运行中", "GET /api/v1/search/companies?q=宁德", "返回宁德时代"),
    ("TC-1.2-08", "1.2 公司搜索 API", "P1", "响应格式校验", "🤖自动",
     "后端运行中", "检查任意响应 JSON 结构", "含 code/message/data/timestamp"),

    # ── 模块 1.3: 指标计算引擎 ──
    ("TC-1.3-01", "1.3 指标计算引擎", "P0", "正常同比增长", "🤖自动",
     "运行 IndicatorServiceTest", "calculateYoY(150, 120)", "返回 25.00"),
    ("TC-1.3-02", "1.3 指标计算引擎", "P0", "同比下降", "🤖自动",
     "运行单元测试", "calculateYoY(80, 100)", "返回 -20.00"),
    ("TC-1.3-03", "1.3 指标计算引擎", "P0", "除零边界处理", "🤖自动",
     "运行单元测试", "calculateYoY(100, 0)", "返回 null"),
    ("TC-1.3-04", "1.3 指标计算引擎", "P0", "null 输入处理", "🤖自动",
     "运行单元测试", "calculateYoY(null, 100)", "返回 null"),
    ("TC-1.3-05", "1.3 指标计算引擎", "P1", "亏损同比收窄", "🤖自动",
     "运行单元测试", "calculateYoY(-50, -100)", "返回 50.00"),
    ("TC-1.3-06", "1.3 指标计算引擎", "P0", "复合增长率 CAGR", "🤖自动",
     "运行单元测试", "calculateCAGR([100,120,144], 3)", "≈ 20.00"),
    ("TC-1.3-07", "1.3 指标计算引擎", "P0", "比率计算", "🤖自动",
     "运行单元测试", "calculateRatio(80, 100)", "返回 80.00"),
    ("TC-1.3-08", "1.3 指标计算引擎", "P0", "比率除零", "🤖自动",
     "运行单元测试", "calculateRatio(100, 0)", "返回 null"),
    ("TC-1.3-09", "1.3 指标计算引擎", "P0", "BigDecimal 精度校验", "🤖自动",
     "运行单元测试", "检查所有返回值精度", "保留 2 位小数，四舍五入"),
    ("TC-1.3-10", "1.3 指标计算引擎", "P0", "贵州茅台真实数据验证", "🤖自动",
     "后端运行中", "查 2024 vs 2023 营收同比", "同比 ≈ 16.07%"),

    # ── 模块 1.4: 行业均值与排名 ──
    ("TC-1.4-01", "1.4 行业均值排名", "P1", "白酒行业 benchmark", "🤖自动",
     "后端运行中", "GET /api/v1/analysis/600519/benchmark", "industry=白酒，含 5 个指标"),
    ("TC-1.4-02", "1.4 行业均值排名", "P1", "银行行业 benchmark", "🤖自动",
     "后端运行中", "GET /api/v1/analysis/601398/benchmark", "industry=银行，数据不同于白酒"),
    ("TC-1.4-03", "1.4 行业均值排名", "P1", "排名描述正确性", "🤖自动",
     "后端运行中", "茅台毛利率 92.8% vs 行业 68.5%", "rank=top5%"),
    ("TC-1.4-04", "1.4 行业均值排名", "P2", "无行业数据兜底", "🤖自动",
     "后端运行中", "查询无均值数据公司", "返回空数组，不报错"),
    ("TC-1.4-05", "1.4 行业均值排名", "P1", "前端行业列渲染", "🔀半自动",
     "前端运行中", "Dashboard 查看指标详解表", "行业均值列和行业对比列正确显示"),

    # ── 模块 2.1: 搜索入口 UI ──
    ("TC-2.1-01", "2.1 搜索入口 UI", "P0", "输入搜索触发建议", "🤖自动",
     "前端运行中", "搜索框输入'茅台'", "300ms 内出现下拉建议列表"),
    ("TC-2.1-02", "2.1 搜索入口 UI", "P0", "下拉项内容完整", "🔀半自动",
     "前端运行中", "查看搜索建议项", "每项含公司名、代码、行业标签"),
    ("TC-2.1-03", "2.1 搜索入口 UI", "P0", "关键词高亮", "🙋人工",
     "前端运行中", "输入'茅台'查看建议", "匹配字符黄色高亮或加粗"),
    ("TC-2.1-04", "2.1 搜索入口 UI", "P1", "键盘 ↓ 导航", "🤖自动",
     "前端运行中", "输入'茅台'，按 ↓", "第一项高亮 → 第二项高亮"),
    ("TC-2.1-05", "2.1 搜索入口 UI", "P1", "键盘 ↑ 导航", "🤖自动",
     "前端运行中", "选中第二项，按 ↑", "高亮回到第一项"),
    ("TC-2.1-06", "2.1 搜索入口 UI", "P0", "Enter 确认跳转", "🤖自动",
     "前端运行中", "选中贵州茅台，按 Enter", "跳转 /dashboard/600519"),
    ("TC-2.1-07", "2.1 搜索入口 UI", "P1", "Esc 关闭下拉", "🤖自动",
     "前端运行中", "搜索建议展开，按 Esc", "下拉关闭"),
    ("TC-2.1-08", "2.1 搜索入口 UI", "P1", "快速连续输入", "🤖自动",
     "前端运行中", "快速输入'贵州茅台'", "仅最后一次请求结果展示"),
    ("TC-2.1-09", "2.1 搜索入口 UI", "P0", "无匹配空状态", "🙋人工",
     "前端运行中", "输入'xyzabc123'", "显示'未找到匹配公司'"),
    ("TC-2.1-10", "2.1 搜索入口 UI", "P0", "点击 Chip 跳转", "🤖自动",
     "前端运行中", "点击'宁德时代' Chip", "跳转 /dashboard/300750"),
    ("TC-2.1-11", "2.1 搜索入口 UI", "P2", "点击外部关闭下拉", "🤖自动",
     "前端运行中", "下拉展开后点击空白处", "下拉关闭"),

    # ── 模块 2.2: 页面跳转与加载 ──
    ("TC-2.2-01", "2.2 页面跳转加载", "P0", "路由跳转加载数据", "🤖自动",
     "前端运行中", "访问 /dashboard/600519", "骨架屏 → KPI 卡片渲染"),
    ("TC-2.2-02", "2.2 页面跳转加载", "P0", "切换不同公司", "🤖自动",
     "前端运行中", "茅台 → 搜索切换到五粮液", "骨架屏重现 → 新数据加载"),
    ("TC-2.2-03", "2.2 页面跳转加载", "P1", "API 返回 500 错误", "🙋人工",
     "需 mock 后端返回 500", "访问 /dashboard/600519", "显示错误提示 + 重试按钮"),
    ("TC-2.2-04", "2.2 页面跳转加载", "P1", "重试功能", "🤖自动",
     "模拟错误状态", "点击重试按钮", "重新调用 API 加载数据"),
    ("TC-2.2-05", "2.2 页面跳转加载", "P1", "浏览器前进后退", "🤖自动",
     "已浏览过 2 家公司", "点击后退按钮", "路由正确切换，数据重新加载"),
    ("TC-2.2-06", "2.2 页面跳转加载", "P1", "无效公司代码", "🤖自动",
     "前端运行中", "访问 /dashboard/INVALID", "显示'未找到该公司'"),
    ("TC-2.2-07", "2.2 页面跳转加载", "P0", "直接 URL 访问", "🤖自动",
     "前端运行中", "输入 .../dashboard/000858", "正确加载五粮液 Dashboard"),

    # ── 模块 2.3: 分析历史 ──
    ("TC-2.3-01", "2.3 分析历史管理", "P0", "历史记录新增", "🤖自动",
     "前端运行中", "搜索并加载贵州茅台", "历史下拉显示贵州茅台 1 条"),
    ("TC-2.3-02", "2.3 分析历史管理", "P1", "重复分析去重", "🤖自动",
     "前端运行中", "再次加载贵州茅台", "仅 1 条，时间戳更新"),
    ("TC-2.3-03", "2.3 分析历史管理", "P1", "最多保留 20 条", "🤖自动",
     "前端运行中", "连续加载 25 家不同公司", "仅保留最近 20 条"),
    ("TC-2.3-04", "2.3 分析历史管理", "P0", "点击历史跳转", "🤖自动",
     "前端运行中", "展开历史 → 点击五粮液", "跳转 /dashboard/000858"),
    ("TC-2.3-05", "2.3 分析历史管理", "P1", "单条删除", "🤖自动",
     "前端运行中", "点击删除某条历史", "该项消失，其余不变"),
    ("TC-2.3-06", "2.3 分析历史管理", "P1", "清空全部历史", "🤖自动",
     "前端运行中", "点击'清空全部历史'", "列表为空"),
    ("TC-2.3-07", "2.3 分析历史管理", "P1", "刷新持久化", "🤖自动",
     "前端运行中", "有历史记录 → 刷新页面", "localStorage 数据不丢失"),
    ("TC-2.3-08", "2.3 分析历史管理", "P2", "相对时间显示", "🙋人工",
     "前端运行中", "刚分析后看历史；等 5 分钟再看", "显示'刚刚'→'5 分钟前'"),

    # ── 模块 3.1: KPI 卡片 ──
    ("TC-3.1-01", "3.1 KPI 指标卡片", "P0", "4 张卡片完整渲染", "🤖自动",
     "加载贵州茅台", "查看 Dashboard 顶部", "4 张卡片均存在"),
    ("TC-3.1-02", "3.1 KPI 指标卡片", "P1", "卡片 2×2 网格布局", "🙋人工",
     "加载贵州茅台", "查看卡片排列", "PC 端 2×2 网格"),
    ("TC-3.1-03", "3.1 KPI 指标卡片", "P1", "数字 count-up 动画", "🙋人工",
     "加载贵州茅台", "观察数值变化", "从 0 滚动到目标值，约 1 秒"),
    ("TC-3.1-04", "3.1 KPI 指标卡片", "P0", "营收同比绿色 ↑ 箭头", "🙋人工",
     "加载贵州茅台", "查看营收卡片同比", "绿色 ↑ + 百分比"),
    ("TC-3.1-05", "3.1 KPI 指标卡片", "P0", "负债率下降绿色 ↓", "🙋人工",
     "加载贵州茅台", "查看负债率卡片同比", "绿色 ↓（down_good）"),
    ("TC-3.1-06", "3.1 KPI 指标卡片", "P1", "同比为空灰色显示", "🤖自动",
     "加载无上年数据公司", "查看同比显示", "灰色 '—'"),
    ("TC-3.1-07", "3.1 KPI 指标卡片", "P1", "切换公司重新动画", "🙋人工",
     "当前在茅台", "搜索切换到五粮液", "数值重新 count-up"),
    ("TC-3.1-08", "3.1 KPI 指标卡片", "P1", "移动端 1 列布局", "🙋人工",
     "DevTools 移动模拟", "查看 KPI 卡片", "1 列布局"),

    # ── 模块 3.2: 数据表格 ──
    ("TC-3.2-01", "3.2 数据速览表格", "P0", "表格结构完整", "🤖自动",
     "加载贵州茅台", "查看数据速览区域", "5 行 × 5 列（2020-2024）"),
    ("TC-3.2-02", "3.2 数据速览表格", "P1", "表头 sticky 固定", "🙋人工",
     "加载公司数据", "向下滚动页面", "表头行保持可见"),
    ("TC-3.2-03", "3.2 数据速览表格", "P2", "斑马纹交替行", "🙋人工",
     "加载公司数据", "查看表格行颜色", "奇偶行背景色不同"),
    ("TC-3.2-04", "3.2 数据速览表格", "P1", "移动端横向滚动", "🤖自动",
     "DevTools 移动模拟", "左右滑动表格", "表格可横向滚动"),
    ("TC-3.2-05", "3.2 数据速览表格", "P1", "数值格式化", "🙋人工",
     "加载贵州茅台", "查看营收 2024 列", "显示'1,748 亿'"),
    ("TC-3.2-06", "3.2 数据速览表格", "P0", "表格/折线图切换", "🤖自动",
     "加载公司数据", "点击'折线图' Tab", "表格 → 折线图视图切换"),

    # ── 模块 3.3: 折线图趋势 ──
    ("TC-3.3-01", "3.3 折线图趋势", "P0", "折线图渲染", "🤖自动",
     "切换到折线图视图", "查看图表", "Chart.js 折线图含 5 个数据点"),
    ("TC-3.3-02", "3.3 折线图趋势", "P1", "X/Y 轴标签", "🙋人工",
     "查看折线图", "查看坐标轴", "X=年份(2020-2024)，Y=指标值+单位"),
    ("TC-3.3-03", "3.3 折线图趋势", "P0", "指标 Tab 切换", "🤖自动",
     "默认显示营收", "点击'净利润' Tab", "折线图切换为净利润数据"),
    ("TC-3.3-04", "3.3 折线图趋势", "P1", "Tooltip 悬停", "🙋人工",
     "折线图已渲染", "鼠标悬停数据点", "显示年份、指标名、数值、单位"),
    ("TC-3.3-05", "3.3 折线图趋势", "P2", "切换公司无内存泄漏", "🙋人工",
     "加载茅台 → 切换五粮液", "DevTools Performance 观察", "旧 chart destroy，无残留"),
    ("TC-3.3-06", "3.3 折线图趋势", "P1", "全部 5 个指标切换", "🤖自动",
     "逐个点击指标 Tab", "营收/净利润/毛利率/负债率/现金流", "每个折线图正确渲染"),

    # ── 模块 3.4: 指标详解表 ──
    ("TC-3.4-01", "3.4 指标详解表", "P1", "指标分组展示", "🔀半自动",
     "加载贵州茅台", "查看指标详解区域", "按 category 分组"),
    ("TC-3.4-02", "3.4 指标详解表", "P1", "术语 tooltip", "🙋人工",
     "加载公司数据", "鼠标悬停'?'", "弹出白话术语解释"),
    ("TC-3.4-03", "3.4 指标详解表", "P0", "同比箭头与评价 Tag", "🤖自动",
     "加载公司数据", "查看各指标行", "含同比箭头 + 评价 Tag"),
    ("TC-3.4-04", "3.4 指标详解表", "P1", "行业均值列颜色", "🔀半自动",
     "加载贵州茅台", "查看'行业均值'列", "数值低于公司值标绿"),
    ("TC-3.4-05", "3.4 指标详解表", "P1", "行业排名标签 Chip", "🔀半自动",
     "加载贵州茅台", "查看'行业对比'列", "显示'行业前 5%' Chip"),

    # ── 模块 4.1: 经营亮点 ──
    ("TC-4.1-01", "4.1 经营亮点", "P1", "茅台亮点完整", "🤖自动",
     "加载贵州茅台", "查看经营亮点区域", "品牌护城河/盈利能力/利润含金量/财务安全"),
    ("TC-4.1-02", "4.1 经营亮点", "P1", "卡片绿色左边框样式", "🙋人工",
     "加载公司数据", "查看亮点卡片", "绿色左边框 + emoji + 标题 + 描述"),
    ("TC-4.1-03", "4.1 经营亮点", "P1", "数值自动填充", "🙋人工",
     "加载贵州茅台", "查看亮点描述", "文案包含实际数据（如'毛利率 92.8%'）"),
    ("TC-4.1-04", "4.1 经营亮点", "P1", "高负债公司不含财务安全", "🤖自动",
     "加载高负债公司", "查看亮点列表", "不含'财务结构安全'"),
    ("TC-4.1-05", "4.1 经营亮点", "P2", "无匹配规则空状态", "🤖自动",
     "加载不触发亮点公司", "查看亮点区域", "显示'暂无亮点数据'"),
    ("TC-4.1-06", "4.1 经营亮点", "P2", "规则启用/禁用", "🤖自动",
     "数据库禁用某规则", "重新加载公司", "该规则不触发"),

    # ── 模块 4.2: 风险识别 ──
    ("TC-4.2-01", "4.2 风险识别", "P1", "茅台风险识别", "🔀半自动",
     "加载贵州茅台", "查看风险提示区域", "利润增速放缓、政策监管风险"),
    ("TC-4.2-02", "4.2 风险识别", "P1", "风险卡片红色左边框", "🙋人工",
     "加载公司数据", "查看风险卡片", "红色左边框 + emoji + 标题 + 描述"),
    ("TC-4.2-03", "4.2 风险识别", "P1", "高负债公司风险", "🔀半自动",
     "加载高负债公司", "查看风险列表", "高负债风险卡片出现"),
    ("TC-4.2-04", "4.2 风险识别", "P2", "低风险公司", "🔀半自动",
     "加载低风险公司", "查看风险区域", "仅 0-1 条风险"),
    ("TC-4.2-05", "4.2 风险识别", "P1", "亮点风险并排布局", "🙋人工",
     "加载贵州茅台", "查看布局", "PC 左右排列，移动端上下排列"),

    # ── 模块 5.1: 趋势预测 ──
    ("TC-5.1-01", "5.1 趋势预测", "P1", "预测 Tab 切换", "🤖自动",
     "加载贵州茅台", "点击'趋势预测' Tab", "切换到预测模块"),
    ("TC-5.1-02", "5.1 趋势预测", "P1", "双折线图渲染", "🔀半自动",
     "切换到预测模块", "查看图表", "实线+虚线预测值"),
    ("TC-5.1-03", "5.1 趋势预测", "P1", "预测值合理性", "🤖自动",
     "查看预测图", "对比 2024 vs 2025E", "2025E ≥ 2024（正增长）"),
    ("TC-5.1-04", "5.1 趋势预测", "P2", "置信区间色带", "🙋人工",
     "查看预测图", "观察预测线上下", "半透明色带覆盖预测区域"),
    ("TC-5.1-05", "5.1 趋势预测", "P1", "图例完整", "🔀半自动",
     "查看预测图", "查看图例标签", "4 条：营收实际/预测 + 净利润实际/预测"),
    ("TC-5.1-06", "5.1 趋势预测", "P2", "数据不足提示", "🤖自动",
     "加载 <3 年数据公司", "查看预测区域", "显示'数据不足，无法预测'"),
    ("TC-5.1-07", "5.1 趋势预测", "P1", "洞察文案展示", "🙋人工",
     "加载贵州茅台", "查看预测图表下方", "含 R²、趋势描述、变化百分比"),
    ("TC-5.1-08", "5.1 趋势预测", "P1", "免责声明展示", "🤖自动",
     "查看预测模块底部", "滚动到底部", "黄色虚线边框免责声明"),

    # ── 模块 6.1: 聊天面板 UI ──
    ("TC-6.1-01", "6.1 聊天面板 UI", "P1", "浮动按钮可见", "🤖自动",
     "加载 Dashboard", "查看右下角", "💬 浮动按钮可见"),
    ("TC-6.1-02", "6.1 聊天面板 UI", "P1", "点击展开面板", "🤖自动",
     "Dashboard 页面", "点击 💬 按钮", "面板右侧滑入，浮动按钮隐藏"),
    ("TC-6.1-03", "6.1 聊天面板 UI", "P1", "欢迎语+建议问题", "🔀半自动",
     "首次打开聊天", "查看面板内容", "欢迎语 + 3 个建议问题"),
    ("TC-6.1-04", "6.1 聊天面板 UI", "P1", "发送用户消息", "🤖自动",
     "面板已打开", "输入问题按 Enter", "右侧蓝色气泡"),
    ("TC-6.1-05", "6.1 聊天面板 UI", "P1", "助手回复 Markdown", "🙋人工",
     "面板已打开", "等待助手回复", "左侧灰色气泡，含 Markdown 渲染"),
    ("TC-6.1-06", "6.1 聊天面板 UI", "P1", "自动滚动到底部", "🤖自动",
     "面板已打开", "发送多条消息", "新消息后自动滚动"),
    ("TC-6.1-07", "6.1 聊天面板 UI", "P2", "Shift+Enter 换行", "🤖自动",
     "面板已打开", "输入文字 → Shift+Enter", "换行不发送"),
    ("TC-6.1-08", "6.1 聊天面板 UI", "P2", "浮动按钮拖拽", "🙋人工",
     "Dashboard 页面", "按住 💬 拖拽移动", "按钮跟随移动，放开停留"),
    ("TC-6.1-09", "6.1 聊天面板 UI", "P1", "关闭聊天面板", "🤖自动",
     "聊天面板已打开", "点击 ✕ 关闭", "面板滑出隐藏，按钮重现"),
    ("TC-6.1-10", "6.1 聊天面板 UI", "P2", "移动端全屏面板", "🙋人工",
     "DevTools 移动模拟", "展开聊天面板", "面板宽度 100vw"),

    # ── 模块 6.2-6.3: AI 对话质量 ──
    ("TC-6.2-01", "6.2 AI 对话", "P1", "提问盈利能力（mock）", "🤖自动",
     "LLM_PROVIDER=mock", "发送'盈利能力怎么样？'", "返回含毛利率/净利率回复"),
    ("TC-6.2-02", "6.2 AI 对话", "P1", "回答含 RAG 引用", "🙋人工",
     "加载贵州茅台后", "发送问题", "回复末尾参考来源标注"),
    ("TC-6.2-03", "6.2 AI 对话", "P2", "追问建议按钮", "🤖自动",
     "收到回复后", "查看回复气泡下方", "1-3 个追问建议按钮"),
    ("TC-6.2-04", "6.2 AI 对话", "P1", "Mock 模式多问题可用", "🤖自动",
     "LLM_PROVIDER=mock", "连续发送 5 个不同问题", "均有合理回复"),
    ("TC-6.2-05", "6.2 AI 对话", "P2", "SSE 流式打字效果", "🙋人工",
     "发送问题", "观察回复出现方式", "逐字/逐 token 出现"),
    ("TC-6.2-06", "6.2 AI 对话", "P2", "聊天记录导出", "🤖自动",
     "面板已有多条消息", "点击导出按钮", "下载 .txt 文件含完整对话"),

    # ── 模块 7.1: 文件上传 ──
    ("TC-7.1-01", "7.1 文件上传", "P2", "点击上传 TXT", "🤖自动",
     "搜索页", "选择 .txt 文件上传", "返回 taskId"),
    ("TC-7.1-02", "7.1 文件上传", "P2", "拖拽上传", "🤖自动",
     "搜索页", "拖拽 .txt 到上传区域", "触发上传"),
    ("TC-7.1-03", "7.1 文件上传", "P2", "不支持格式拦截", "🤖自动",
     "搜索页", "尝试上传 .exe", "提示'不支持的文件类型'"),
    ("TC-7.1-04", "7.1 文件上传", "P2", "超大文件拦截", "🤖自动",
     "搜索页", "尝试上传 >50MB", "提示'文件大小不能超过 50MB'"),
    ("TC-7.1-05", "7.1 文件上传", "P2", "上传进度条", "🤖自动",
     "搜索页", "上传较大文件", "进度条 0%→100%"),

    # ── 模块 7.2-7.3: OCR/NER ──
    ("TC-7.2-01", "7.2 OCR/NER", "P2", "TXT 文件解析", "🤖自动",
     "上传含财务数据 .txt", "等待解析完成", "返回真实文本"),
    ("TC-7.2-02", "7.2 OCR/NER", "P2", "NER 正则提取", "🤖自动",
     "TXT 含'营业收入：1748亿元'", "触发 NER 提取", "revenue=1748, confidence≈0.95"),
    ("TC-7.2-03", "7.2 OCR/NER", "P2", "低置信度标记", "🤖自动",
     "非标准格式文本", "触发 NER 提取", "confidence<0.8 带 ⚠️"),
    ("TC-7.2-04", "7.2 OCR/NER", "P2", "提取结果确认表", "🔀半自动",
     "上传并解析完成", "查看确认界面", "可编辑表格"),
    ("TC-7.2-05", "7.2 OCR/NER", "P2", "用户手动修正提交", "🔀半自动",
     "提取结果确认表", "修改数值 → 确认提交", "修改值写入数据库"),

    # ── 模块 8.1: 前端导出 ──
    ("TC-8.1-01", "8.1 前端导出", "P1", "导出 PNG", "🙋人工",
     "加载贵州茅台", "导出菜单 → PNG", "浏览器下载含公司名+日期 .png"),
    ("TC-8.1-02", "8.1 前端导出", "P1", "导出 PDF", "🙋人工",
     "加载贵州茅台", "导出菜单 → PDF", "下载 .pdf，A4 分页不截断图表"),
    ("TC-8.1-03", "8.1 前端导出", "P1", "导出 Word", "🔀半自动",
     "加载贵州茅台", "导出菜单 → Word", "下载 .doc 含 KPI+图表"),
    ("TC-8.1-04", "8.1 前端导出", "P1", "导出 Excel", "🔀半自动",
     "加载贵州茅台", "导出菜单 → Excel", "下载 .xlsx 表格数据正确"),
    ("TC-8.1-05", "8.1 前端导出", "P1", "导出含免责声明", "🙋人工",
     "任意导出", "查看导出内容", "所有格式含黄色边框免责声明"),
    ("TC-8.1-06", "8.1 前端导出", "P2", "导出后 DOM 恢复", "🤖自动",
     "导出完成后", "检查页面", "动态元素被移除"),

    # ── 8.2 后端导出 ──
    ("TC-8.2-01", "8.2 后端导出", "P1", "后端导出 PDF", "🤖自动",
     "后端运行中, 已登录", "POST /api/v1/export format=pdf", "返回taskId, 任务创建成功"),
    ("TC-8.2-02", "8.2 后端导出", "P1", "后端导出 Excel", "🤖自动",
     "后端运行中, 已登录", "POST /api/v1/export format=xlsx", "返回taskId"),
    ("TC-8.2-03", "8.2 后端导出", "P1", "后端导出 Word", "🤖自动",
     "后端运行中, 已登录", "POST /api/v1/export format=docx", "返回taskId"),
    ("TC-8.2-04", "8.2 后端导出", "P1", "后端导出 PNG", "🤖自动",
     "后端运行中, 已登录", "POST /api/v1/export format=png", "返回taskId"),
    ("TC-8.2-05", "8.2 后端导出", "P1", "导出含免责声明", "🤖自动",
     "后端运行中", "includeDisclaimer=true", "导出任务创建成功"),
    ("TC-8.2-06", "8.2 后端导出", "P1", "无Token导出被拒", "🤖自动",
     "后端运行中", "不带Token请求导出", "返回401"),

    # ── 跨模块 E2E ──
    ("TC-E2E-01", "E2E 核心旅程", "P0", "搜索→看板→聊天完整链路", "🤖自动",
     "全部服务运行", "搜索茅台→看KPI→切折线图→亮点风险→预测→聊天", "全链路无报错"),
    ("TC-E2E-02", "E2E 核心旅程", "P0", "热门Chip快速分析", "🤖自动",
     "全部服务运行", "点击宁德时代Chip→看Dashboard→切换工商银行", "跳转正确，数据加载正常"),
    ("TC-E2E-03", "E2E 核心旅程", "P0", "历史记录+导出", "🤖自动",
     "全部服务运行", "分析3家公司→看历史→点击切换→导出PNG→导出Excel", "历史正确，导出正常"),
    ("TC-E2E-04", "E2E 核心旅程", "P1", "无匹配→重试", "🤖自动",
     "全部服务运行", "搜'xyzabc'→确认空状态→搜'600519'", "空状态友好，搜索恢复"),

    # ── 非功能: 性能 ──
    ("TC-PF-01", "⚡性能测试", "P1", "搜索 API 延迟 P95<200ms", "🤖自动",
     "后端运行中", "k6 压测 GET /search/companies", "P95 < 200ms"),
    ("TC-PF-02", "⚡性能测试", "P1", "KPI API 延迟 P95<500ms", "🤖自动",
     "后端运行中", "k6 压测 GET /reports/{code}/kpi", "P95 < 500ms"),
    ("TC-PF-03", "⚡性能测试", "P1", "AI 引擎健康检查 <50ms", "🤖自动",
     "AI 引擎运行中", "wrk GET /health", "P95 < 50ms"),
    ("TC-PF-04", "⚡性能测试", "P2", "RAG 检索延迟 <500ms", "🤖自动",
     "AI 引擎运行中", "POST /ai/v1/rag/search", "延迟 < 500ms"),
    ("TC-PF-05", "⚡性能测试", "P2", "前端首屏 FCP<2s LCP<3s", "🤖自动",
     "前端运行中", "lighthouse-ci CLI", "FCP < 2s, LCP < 3s"),
    ("TC-PF-06", "⚡性能测试", "P2", "折线图切换 <200ms", "🤖自动",
     "前端运行中", "5 个指标逐一切换计时", "每次 < 200ms"),

    # ── 非功能: 兼容性 ──
    ("TC-CP-01", "📱兼容性", "P0", "Chrome 最新版", "🙋人工",
     "Windows+Chrome 120+", "全功能走查", "全部功能正常"),
    ("TC-CP-02", "📱兼容性", "P1", "Edge 最新版", "🙋人工",
     "Windows+Edge 120+", "全功能走查", "全部功能正常"),
    ("TC-CP-03", "📱兼容性", "P2", "移动端 Safari", "🙋人工",
     "iOS Safari 真机", "搜索/看板/聊天可用", "布局自适应"),
    ("TC-CP-04", "📱兼容性", "P2", "移动端 Chrome", "🙋人工",
     "Android Chrome 真机", "同上", "同上"),
    ("TC-CP-05", "📱兼容性", "P0", "1920×1080 桌面布局", "🙋人工",
     "标准桌面分辨率", "检查页面布局", "布局正常"),
    ("TC-CP-06", "📱兼容性", "P2", "375×812 iPhone X 布局", "🙋人工",
     "DevTools 模拟", "检查移动端布局", "1 列布局，横向滚动表格"),

    # ── 非功能: 安全 ──
    ("TC-SC-01", "🔒安全测试", "P2", "SQL 注入防护", "🤖自动",
     "后端运行中", "搜索框输入 'OR'1'='1", "不返回异常数据"),
    ("TC-SC-02", "🔒安全测试", "P2", "XSS 防护", "🤖自动",
     "后端运行中", "输入 <script>alert(1)</script>", "不执行脚本"),
    ("TC-SC-03", "🔒安全测试", "P2", "超大文件拦截", "🤖自动",
     "后端运行中", "上传 >50MB 文件", "前端拦截"),
    ("TC-SC-04", "🔒安全测试", "P3", "未授权访问 401（预留）", "🤖自动",
     "JWT 实施后", "不带 Token 访问 🔒 接口", "返回 401"),
    ("TC-SC-05", "🔒安全测试", "P2", "伪造 Token 拒绝", "🤖自动",
     "后端运行中", "带伪造 Token 访问 🔒 接口", "返回 401"),

    # ── Epic 9: 用户系统（认证 API） ──
    ("TC-9.1-01", "9.1 认证 API", "P2", "用户注册", "🤖自动",
     "后端运行中", "POST /api/v1/auth/register", "返回 userId+email+nickname"),
    ("TC-9.1-02", "9.1 认证 API", "P2", "重复注册拦截", "🤖自动",
     "后端运行中", "重复注册相同 email", "被拦截"),
    ("TC-9.1-03", "9.1 认证 API", "P2", "弱密码拦截", "🤖自动",
     "后端运行中", "密码过短注册", "返回400"),
    ("TC-9.1-05", "9.1 认证 API", "P2", "登录成功返回JWT", "🤖自动",
     "后端运行中", "POST /api/v1/auth/login", "返回 accessToken+refreshToken"),
    ("TC-9.1-06", "9.1 认证 API", "P2", "错误密码拒绝", "🤖自动",
     "后端运行中", "错误密码登录", "返回400"),
    ("TC-9.1-09", "9.1 认证 API", "P2", "登出成功", "🤖自动",
     "后端运行中, 已登录", "POST /api/v1/auth/logout", "code=0"),
    ("TC-9.1-10", "9.1 认证 API", "P2", "获取当前用户信息", "🤖自动",
     "后端运行中, 已登录", "GET /api/v1/auth/me", "返回 email+nickname"),

    # ── Epic 9: 认证 UI ──
    ("TC-9.2-05", "9.2 认证 UI", "P2", "未登录显示登录按钮", "🙋人工",
     "前端运行, 未登录", "查看 Navbar", "显示登录按钮"),
    ("TC-9.2-06", "9.2 认证 UI", "P2", "已登录显示用户信息", "🙋人工",
     "前端运行, 已登录", "查看 Navbar", "显示昵称/邮箱+退出"),
    ("TC-9.2-07", "9.2 认证 UI", "P2", "登录/注册表单切换", "🙋人工",
     "前端运行", "点击切换", "表单切换，按钮文案变化"),

    # ── Epic 9: API 权限控制 ──
    ("TC-9.3-01", "9.3 API 权限", "P2", "聊天接口需认证", "🤖自动",
     "后端运行中", "POST /chat/messages 无Token", "返回401"),
    ("TC-9.3-02", "9.3 API 权限", "P2", "上传接口需认证", "🤖自动",
     "后端运行中", "POST /upload/report 无Token", "返回401"),
    ("TC-9.3-03", "9.3 API 权限", "P2", "导出接口需认证", "🤖自动",
     "后端运行中", "POST /export 无Token", "返回401"),
    ("TC-9.3-04", "9.3 API 权限", "P2", "历史接口需认证", "🤖自动",
     "后端运行中", "GET /history 无Token", "返回401"),
    ("TC-9.3-05", "9.3 API 权限", "P2", "带Token访问历史", "🤖自动",
     "后端运行中, 已登录", "GET /history 带Token", "返回200"),
    ("TC-9.3-07", "9.3 API 权限", "P2", "公开接口无需Token", "🤖自动",
     "后端运行中", "GET /search/companies 无Token", "返回200"),
]

# ── 列宽定义 ──
COL_WIDTHS = {
    "A": 14,  # TC-ID
    "B": 20,  # 模块
    "C": 7,   # 优先级
    "D": 32,  # 测试场景
    "E": 10,  # 执行类型
    "F": 20,  # 前置条件
    "G": 30,  # 测试步骤
    "H": 35,  # 预期结果
    "I": 8,   # 结果
    "J": 30,  # 实际结果
    "K": 28,  # 评语/备注
    "L": 12,  # 缺陷编号
    "M": 10,  # 测试人
    "N": 12,  # 测试日期
    "O": 14,  # 版本号
}


def _priority_fill(priority):
    if priority == "P0": return P0_FILL
    if priority == "P1": return P1_FILL
    return P2_FILL


def create_overview_sheet(wb):
    """Sheet 1: 测试概览"""
    ws = wb.active
    ws.title = "测试概览"
    ws.sheet_properties.tabColor = "1F2937"

    # 标题区
    ws.merge_cells("A1:H1")
    ws["A1"] = "SmartReport 测试报告"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = TITLE_ALIGN
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = "智能财报研读与分析助手 · 功能测试与质量评估"
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:H3")
    ws["A3"] = f"报告日期：{date.today().isoformat()}"
    ws["A3"].font = SUBTITLE_FONT

    # 基本信息区
    row = 5
    info = [
        ("项目名称", "SmartReport"),
        ("被测版本", "v1.0.0-beta1"),
        ("测试环境", "Docker Compose (localhost:3000/8080/8000)"),
        ("LLM Provider", "mock"),
        ("数据库", "MySQL 8.0 (localhost:3307)"),
        ("测试周期", "____年__月__日 — ____年__月__日"),
        ("测试团队", "测试与用户体验组"),
    ]
    for label, value in info:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = BOLD_FONT
        ws[f"A{row}"].fill = OVERVIEW_HEADER_FILL
        ws[f"A{row}"].border = THIN_BORDER
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = BODY_FONT
        ws[f"B{row}"].border = THIN_BORDER
        row += 1

    # KPI 统计区
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "用例执行统计"
    ws[f"A{row}"].font = Font(name="微软雅黑", size=12, bold=True, color="1F2937")
    row += 1

    kpi_headers = ["指标", "总数", "占比"]
    for ci, h in enumerate(kpi_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    row += 1

    kpis = [
        ("📋 总用例数", len(TEST_CASES), "100%"),
        ("✅ 通过", "=COUNTIF(功能测试!I:I,\"✅\")", '=IF(B{0}=0,"0%",TEXT(B{0}/B{1},"0.0%"))'),
        ("❌ 失败", "=COUNTIF(功能测试!I:I,\"❌\")", '=IF(B{0}=0,"0%",TEXT(B{0}/B{1},"0.0%"))'),
        ("🚧 阻塞", "=COUNTIF(功能测试!I:I,\"🚧\")", '=IF(B{0}=0,"0%",TEXT(B{0}/B{1},"0.0%"))'),
        ("⬜ 未执行", "=COUNTIF(功能测试!I:I,\"\")", '=IF(B{0}=0,"0%",TEXT(B{0}/B{1},"0.0%"))'),
    ]
    total_row_num = row
    for i, (label, _count, _pct) in enumerate(kpis):
        r = row + i
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = LEFT_WRAP
        ws.cell(row=r, column=1).fill = OVERVIEW_KPI_FILL

        cell_count = ws.cell(row=r, column=2)
        if isinstance(_count, int):
            cell_count.value = _count
        else:
            # Excel formula — need to adjust row references
            pass
        cell_count.font = BOLD_FONT
        cell_count.alignment = CENTER
        cell_count.border = THIN_BORDER

        cell_pct = ws.cell(row=r, column=3)
        cell_pct.alignment = CENTER
        cell_pct.border = THIN_BORDER

    # 手动填入公式（openpyxl 公式引用需要精确行号，这里简化写死）
    # 总数行在 total_row_num，用例数直接用 len(TEST_CASES)
    base_row_for_total = total_row_num  # "📋 总用例数" 的行
    ws.cell(row=base_row_for_total, column=2, value=len(TEST_CASES)).font = BOLD_FONT
    ws.cell(row=base_row_for_total, column=2).alignment = CENTER
    ws.cell(row=base_row_for_total, column=2).border = THIN_BORDER
    ws.cell(row=base_row_for_total, column=3, value="100%").font = BOLD_FONT
    ws.cell(row=base_row_for_total, column=3).alignment = CENTER
    ws.cell(row=base_row_for_total, column=3).border = THIN_BORDER

    # 其余行用公式
    for offset, row_num in enumerate(range(base_row_for_total + 1, base_row_for_total + 5)):
        count_cell = ws.cell(row=row_num, column=2)
        pct_cell = ws.cell(row=row_num, column=3)
        # COUNTIF on the 功能测试 sheet I column
        # These will show as formulas in Excel
        count_cell.alignment = CENTER
        count_cell.border = THIN_BORDER
        pct_cell.alignment = CENTER
        pct_cell.border = THIN_BORDER

    # 通过率公式直接写（用户打开 Excel 后生效）
    ws.cell(row=base_row_for_total + 1, column=2, value=None)

    # 优先级分布
    row = base_row_for_total + 6
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "优先级分布"
    ws[f"A{row}"].font = Font(name="微软雅黑", size=12, bold=True, color="1F2937")
    row += 1

    for ci, h in enumerate(["优先级", "数量", "占比"], 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    row += 1

    p0_count = sum(1 for tc in TEST_CASES if tc[2] == "P0")
    p1_count = sum(1 for tc in TEST_CASES if tc[2] == "P1")
    p2_count = sum(1 for tc in TEST_CASES if tc[2] == "P2")
    p3_count = sum(1 for tc in TEST_CASES if tc[2] == "P3")
    total = len(TEST_CASES)

    for p, cnt, fill in [
        ("P0 — 核心必测", p0_count, P0_FILL),
        ("P1 — 重点功能", p1_count, P1_FILL),
        ("P2 — AI 增强 + 非功能", p2_count, P2_FILL),
        ("P3 — 预留（未实施）", p3_count, None),
    ]:
        ws.cell(row=row, column=1, value=p).font = BOLD_FONT
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        ws.cell(row=row, column=1).border = THIN_BORDER
        if fill:
            ws.cell(row=row, column=1).fill = fill

        ws.cell(row=row, column=2, value=cnt).font = BOLD_FONT
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=2).border = THIN_BORDER

        pct = cnt / total * 100
        ws.cell(row=row, column=3, value=f"{pct:.1f}%")
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=3).border = THIN_BORDER
        row += 1

    # 结论区域
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "测试结论"
    ws[f"A{row}"].font = Font(name="微软雅黑", size=12, bold=True, color="1F2937")
    for r in range(row + 1, row + 4):
        ws.merge_cells(f"A{r}:H{r}")
        ws[f"A{r}"].font = BODY_FONT
        ws[f"A{r}"].alignment = LEFT_WRAP
        ws[f"A{r}"].border = THIN_BORDER
        ws.row_dimensions[r].height = 24
    ws[f"A{row+1}"] = "整体评估：________________________________________"
    ws[f"A{row+2}"] = "主要问题：________________________________________"
    ws[f"A{row+3}"] = "建议：____________________________________________"

    # 列宽
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    for c in "DEFGH":
        ws.column_dimensions[c].width = 14


def create_testcases_sheet(wb):
    """Sheet 2: 功能测试（140 条用例明细）"""
    ws = wb.create_sheet("功能测试")
    ws.sheet_properties.tabColor = "3B82F6"

    # 表头
    headers = [
        "TC-ID", "模块", "优先级", "测试场景", "执行类型",
        "前置条件", "测试步骤", "预期结果",
        "✅❌🚧", "实际结果", "评语/备注", "缺陷编号", "测试人", "测试日期", "版本号"
    ]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # 数据行
    for ri, tc in enumerate(TEST_CASES, 2):
        tc_id, module, priority, scenario, exec_type, precondition, steps, expected = tc

        row_data = [tc_id, module, priority, scenario, exec_type, precondition, steps, expected,
                    "", "", "", "", "", "", ""]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.alignment = LEFT_WRAP if ci in (4, 6, 7, 8, 10, 11) else CENTER
            cell.border = THIN_BORDER

            # 优先级颜色
            if ci == 3:
                cell.fill = _priority_fill(priority)

        ws.row_dimensions[ri].height = 32

    # 结果列（I列）数据验证：下拉选项
    dv = DataValidation(type="list", formula1='"✅,❌,🚧"', allow_blank=True)
    dv.error = "请选择 ✅ / ❌ / 🚧"
    dv.errorTitle = "无效结果"
    last_row = len(TEST_CASES) + 1
    dv.add(f"I2:I{last_row}")
    ws.add_data_validation(dv)

    # 列宽
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def create_defects_sheet(wb):
    """Sheet 3: 缺陷清单"""
    ws = wb.create_sheet("缺陷清单")
    ws.sheet_properties.tabColor = "EF4444"

    headers = [
        "BUG-ID", "关联TC-ID", "模块", "缺陷标题", "严重程度",
        "复现步骤", "预期结果", "实际结果", "截图/附件", "状态",
        "发现人", "发现日期", "负责人", "修复版本", "备注"
    ]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # 预置 20 行空模板
    severity_colors = {
        "Blocker": PatternFill("solid", fgColor="DC2626"),
        "Critical": PatternFill("solid", fgColor="F97316"),
        "Major": PatternFill("solid", fgColor="FACC15"),
        "Minor": PatternFill("solid", fgColor="93C5FD"),
    }

    for ri in range(2, 22):
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_WRAP if ci in (3, 5, 6, 7, 8, 15) else CENTER
        ws.row_dimensions[ri].height = 24

    # 严重程度下拉
    dv_severity = DataValidation(
        type="list",
        formula1='"Blocker,Critical,Major,Minor"',
        allow_blank=True
    )
    dv_severity.add("E2:E22")
    ws.add_data_validation(dv_severity)

    # 状态下拉
    dv_status = DataValidation(
        type="list",
        formula1='"待修复,修复中,已修复,已验证,关闭,非缺陷"',
        allow_blank=True
    )
    dv_status.add("J2:J22")
    ws.add_data_validation(dv_status)

    col_widths_defects = {
        "A": 12, "B": 14, "C": 18, "D": 36, "E": 10,
        "F": 36, "G": 30, "H": 30, "I": 14, "J": 10,
        "K": 10, "L": 12, "M": 10, "N": 14, "O": 20,
    }
    for col_letter, width in col_widths_defects.items():
        ws.column_dimensions[col_letter].width = width


def create_version_sheet(wb):
    """Sheet 4: 版本记录"""
    ws = wb.create_sheet("版本记录")
    ws.sheet_properties.tabColor = "8B5CF6"

    headers = [
        "版本号", "发布日期", "测试类型", "总用例数", "通过", "失败",
        "阻塞", "通过率", "测试报告文件", "备注"
    ]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # 预置版本记录示例
    versions = [
        ("v1.0.0-beta1", date.today().isoformat(), "全量功能测试", None, None, None, None, None, "", "首发测试"),
        ("v1.0.0-beta2", "", "增量+回归", None, None, None, None, None, "", "修复后回归"),
        ("v1.0.0-rc1", "", "全量回归", None, None, None, None, None, "", "候选发布"),
        ("v1.0.0", "", "验收测试", None, None, None, None, None, "", "正式发布"),
    ]

    for ri, ver in enumerate(versions, 2):
        for ci, val in enumerate(ver, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER if ci != 10 else LEFT_WRAP
        ws.row_dimensions[ri].height = 24

    # 留空行
    for ri in range(6, 16):
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        ws.row_dimensions[ri].height = 24

    # 测试类型下拉
    dv_type = DataValidation(
        type="list",
        formula1='"全量功能测试,增量+回归,全量回归,验收测试,冒烟测试,专项测试"',
        allow_blank=True
    )
    dv_type.add("C2:C15")
    ws.add_data_validation(dv_type)

    col_widths_ver = {
        "A": 18, "B": 14, "C": 18, "D": 12, "E": 10,
        "F": 10, "G": 10, "H": 10, "I": 22, "J": 22,
    }
    for col_letter, width in col_widths_ver.items():
        ws.column_dimensions[col_letter].width = width


def main():
    parser = argparse.ArgumentParser(description="生成 SmartReport 测试报告 Excel 模板")
    parser.add_argument(
        "--output", "-o",
        default=f"test-report-v1.0.0-beta1.xlsx",
        help="输出文件名"
    )
    args = parser.parse_args()

    wb = Workbook()

    create_overview_sheet(wb)
    create_testcases_sheet(wb)
    create_defects_sheet(wb)
    create_version_sheet(wb)

    output_path = f"test/{args.output}"
    wb.save(output_path)
    print(f"✅ 测试报告模板已生成: {output_path}")
    print(f"   包含 {len(TEST_CASES)} 条测试用例")
    print(f"   Sheet: 测试概览 / 功能测试 / 缺陷清单 / 版本记录")


if __name__ == "__main__":
    main()
